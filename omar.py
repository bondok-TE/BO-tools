from openpyxl import load_workbook
import pandas as pd
wb_list = load_workbook('D:/scripts/nexus_output.xlsx')
sheet = wb_list.active

# D --> name 4
# G --> solution or action 7
# H --> risk factor
# data = {risk_factor:[{name1:solution1},{name2:solution2}]}
# data = {H1: {D1:G1,D2:G2},{H2: {D3:G3,D4:G4}


data = []
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
  
    ls = []
    for cell in row:
        ls.append(cell.value)
    data.append(ls)

desired_data = []
for row in data:
    desired_data.append([row[7],row[3],row[6]])



df = pd.DataFrame(desired_data,columns=['Criticality', 'Vulnerability Name','Action'])

df.to_excel('D:/scripts/my_output.xlsx', index=False)
