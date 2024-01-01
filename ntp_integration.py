
import json
import pandas as pd
import pprint as pp
from openpyxl import load_workbook,Workbook


def dct_ntp_association(ntp_association):
    ntp_association_val = ntp_association['output'][0]['data'].split('\n')
    ntp_association_val.pop(1)
    ntp_association_val = list(map(str.split,ntp_association_val))
    response = {"data":[]}
    for i in range(1,len(ntp_association_val)):
        ntp_association_val_dct_temp = {}
        for ind,val in enumerate(ntp_association_val[0]):
            ntp_association_val_dct_temp[val] = (ntp_association_val[i][ind])
        response['data'].append(ntp_association_val_dct_temp)
        

    # pp.pprint(ntp_association_val_dct_1)
    # print("===========")
    # pp.pprint(ntp_association_val_dct_2)
    # print("===========")
    # pp.pprint(response)


    return response

def manipulate_response(response,ip,hostname):
    # keys to include from the 
    
    # keys_to_include = [
    # "ntp-status.remote-server.hostname",
    # "ntp-status.remote-server.device-ip",
    # "ntp-status.remote-server.server-address",
    # "ntp-status.remote-server.status",
    # ]

    desired_order = [
        "hostname",
        "device-ip",
        "remote",
        "refid",
        "st",
        "t",
        "when",
        "poll",
        "reach",
        "delay",
        "offset",
        "jitter"
    ]

    # add data to the response suitable for viewing in excel
    # pp.pprint(response)
    response['data'][0]["hostname"] = hostname
    response['data'][0]["device-ip"] = ip
    # pp.pprint(response)
    df = pd.json_normalize(response,record_path=['data'],errors="ignore")

    # Reorder columns based on desired_order list

    df = df[desired_order]
    # print(df)
    return df


def create_excel():
    workbook = Workbook()
    workbook.save("ntp_integrations.xlsx")
    file_path = 'ntp_integrations.xlsx'
    wb = load_workbook(file_path)
    sheet = wb.active
    first_row = [
        "hostname",
        "device-ip",
        "remote",
        "refid",
        "st",
        "t",
        "when",
        "poll",
        "reach",
        "delay",
        "offset",
        "jitter"
    ]
    sheet.append(first_row)
    return [wb,sheet]

def write_to_excel(wb_sheet,df):
    wb = wb_sheet[0]
    sheet = wb_sheet[1]
    data_values = df.values.tolist()
    for row in data_values:
        sheet.append(row)
    wb.save("ntp_integrations.xlsx")